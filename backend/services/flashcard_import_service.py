"""
Server-side parsers and Gemini wrappers for flashcard import.

Pure functions only — no Supabase or HTTP coupling beyond the Gemini
client. Each function returns plain dicts the routes can serialize.
"""
from __future__ import annotations

import base64
import io
import json
import logging
import math
import os
import re
import sqlite3
import time
import zipfile
from pathlib import Path
from typing import TypedDict

import httpx
from bs4 import BeautifulSoup
from Levenshtein import distance as _levenshtein
from pydantic_ai.exceptions import ModelHTTPError, UnexpectedModelBehavior

from db.connection import table
from agents._run import run_agent_sync
from agents.flashcard import flashcard_agent
from agents.usage import record_agent_usage
from services import extraction_service

logger = logging.getLogger(__name__)


class Card(TypedDict):
    front: str
    back: str


class QuizletBlocked(Exception):
    """Raised when scrape_quizlet_url cannot extract cards."""


_PROMPTS_DIR = Path(__file__).parent.parent / "prompts"


def _load_prompt(name: str) -> str:
    return (_PROMPTS_DIR / name).read_text(encoding="utf-8")


def _normalize(text: str) -> str:
    return re.sub(r"[^\w\s]", "", text).strip().lower()


def dedup_against_existing(
    user_id: str,
    offering_id: str | None,
    cards: list[Card],
    topic: str | None = None,
) -> tuple[list[Card], list[Card]]:
    """Return (keep, skipped) where skipped have a near-duplicate front
    (Levenshtein <= 3 on normalized front) among the user's existing cards
    in the same offering (or topic when offering_id is None).

    Flashcards key on the course offering after migration 0025 (the link column
    was renamed course_id -> offering_id)."""
    filters = {"user_id": f"eq.{user_id}"}
    if offering_id:
        filters["offering_id"] = f"eq.{offering_id}"
    elif topic:
        filters["topic"] = f"eq.{topic}"

    existing = table("flashcards").select("front", filters=filters) or []
    existing_norm = [_normalize(r.get("front", "")) for r in existing]

    keep: list[Card] = []
    skipped: list[Card] = []
    for card in cards:
        norm = _normalize(card.get("front", ""))
        is_dup = any(_levenshtein(norm, e) <= 3 for e in existing_norm if e)
        (skipped if is_dup else keep).append(card)
    return keep, skipped


_RATE_WINDOW_SEC = 60
_RATE_LIMIT = 5
_rate_state: dict[str, list[float]] = {}


def check_rate_limit(user_id: str) -> int | None:
    """Returns None if call allowed, else seconds until allowed again."""
    now = time.time()
    bucket = [t for t in _rate_state.get(user_id, []) if now - t < _RATE_WINDOW_SEC]
    if len(bucket) >= _RATE_LIMIT:
        # Seconds until the oldest call in the window ages out (freeing a slot),
        # rounded up. ceil keeps a sub-second remainder from reporting 0 and —
        # unlike the old `int(...) + 1` — never overshoots to 61 when the calls
        # land in the same clock tick (elapsed == 0). The min() cap additionally
        # pins the ceiling when `now` reads BEFORE bucket[0] (NTP step / coarse
        # Windows timers, #346), so the result is always in [1, _RATE_WINDOW_SEC].
        retry = min(_RATE_WINDOW_SEC, math.ceil(_RATE_WINDOW_SEC - (now - bucket[0])))
        _rate_state[user_id] = bucket
        return retry
    bucket.append(now)
    _rate_state[user_id] = bucket
    return None


def parse_xlsx(file_bytes: bytes) -> list[Card]:
    """Read a .xlsx workbook, treating col A as front and col B as back."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    cards: list[Card] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        front = (str(row[0]) if len(row) > 0 and row[0] is not None else "").strip()
        back = (str(row[1]) if len(row) > 1 and row[1] is not None else "").strip()
        if front and back:
            cards.append({"front": front, "back": back})
    return cards


def _strip_html(text: str) -> str:
    soup = BeautifulSoup(text or "", "html.parser")
    return re.sub(r"\s+", " ", soup.get_text(" ")).strip()


def parse_anki_apkg(file_bytes: bytes) -> list[Card]:
    """Extract notes from an Anki .apkg (zip with a SQLite collection.anki2)."""
    import tempfile

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            names = z.namelist()
            if "collection.anki2" not in names:
                raise ValueError("Anki package is missing collection.anki2")
            db_bytes = z.read("collection.anki2")
    except zipfile.BadZipFile as e:
        raise ValueError(f"Anki file is not a valid .apkg: {e}")

    with tempfile.NamedTemporaryFile(suffix=".anki2", delete=False) as tmp:
        tmp.write(db_bytes)
        path = tmp.name

    try:
        conn = sqlite3.connect(path)
        rows = conn.execute("SELECT flds FROM notes").fetchall()
        conn.close()
    finally:
        os.unlink(path)

    cards: list[Card] = []
    for (flds,) in rows:
        if not flds:
            continue
        parts = flds.split("\x1f")
        if len(parts) < 2:
            continue
        front = _strip_html(parts[0])
        back = _strip_html(parts[1])
        if front and back:
            cards.append({"front": front, "back": back})
    return cards


_QUIZLET_PAYLOAD_RE = re.compile(
    r'window\.Quizlet\["setPageData"\]\s*=\s*(\{.*?\});',
    re.DOTALL,
)


def scrape_quizlet_url(url: str) -> list[Card]:
    """Best-effort fetch of a public Quizlet set. Raises QuizletBlocked on
    bot wall, login redirect, or unparseable payload."""
    try:
        resp = httpx.get(
            url,
            timeout=15.0,
            headers={"User-Agent": "Mozilla/5.0 (Sapling flashcard import)"},
            follow_redirects=True,
        )
    except httpx.HTTPError as e:
        raise QuizletBlocked(f"Could not reach Quizlet: {e}")

    if resp.status_code != 200:
        raise QuizletBlocked(f"Quizlet returned status {resp.status_code}")

    match = _QUIZLET_PAYLOAD_RE.search(resp.text)
    if not match:
        raise QuizletBlocked(
            "Couldn't extract cards from this URL. Quizlet may be blocking "
            "scrapers — try exporting the set and pasting the text instead."
        )

    try:
        data = json.loads(match.group(1))
    except json.JSONDecodeError as e:
        raise QuizletBlocked(f"Quizlet payload was not valid JSON: {e}")

    terms = (data.get("set") or {}).get("terms") or []
    cards: list[Card] = []
    for t in terms:
        front = (t.get("word") or "").strip()
        back = (t.get("definition") or "").strip()
        if front and back:
            cards.append({"front": front, "back": back})
    if not cards:
        raise QuizletBlocked("No terms found in the Quizlet payload.")
    return cards


def _guess_content_type(filename: str) -> str:
    lower = filename.lower()
    if lower.endswith(".png"):
        return "image/png"
    if lower.endswith(".jpg") or lower.endswith(".jpeg"):
        return "image/jpeg"
    if lower.endswith(".gif"):
        return "image/gif"
    if lower.endswith(".webp"):
        return "image/webp"
    return "image/png"


# Transient provider statuses the legacy call_gemini path retried on. 429 =
# rate limit, 5xx = upstream/transport hiccups.
_TRANSIENT_STATUSES = frozenset({429, 500, 502, 503, 504})
# Mirror the old call_gemini(retries=1): one retry with a ~2s backoff.
_AGENT_RETRIES = 1


def _run_flashcard_agent(prompt: str) -> list[Card]:
    """Run the flashcard agent on a rendered prompt and return front/back dicts.

    Filters out cards missing either side. Error contract (unified across the
    generate + import seams):

    * A model whose output can't satisfy the ``Flashcards`` schema is treated as
      "bad LLM output" and degrades to ``[]`` — logged, never silent. This
      matches the import seams' long-standing bad-JSON ``[]`` resilience; note
      it is a deliberate contract change for ``/generate``, whose legacy path
      *raised* (surfacing a 502) on malformed output.
    * Transient provider errors (HTTP 429/5xx) are retried once with a ~2s
      backoff, then re-raised.
    * Every other failure — transport, missing GEMINI_API_KEY,
      ``run_agent_sync`` misuse from a running event loop, etc. — propagates so
      the route surfaces a retryable 502 instead of a misleading 200 with zero
      cards.
    """
    result = None
    for attempt in range(_AGENT_RETRIES + 1):
        try:
            result = record_agent_usage(
                run_agent_sync(flashcard_agent.run(prompt)),
                feature="flashcard", task="flashcard",
            )
            break
        except UnexpectedModelBehavior:
            logger.exception("flashcard agent produced unusable output; degrading to []")
            return []
        except ModelHTTPError as e:
            if attempt < _AGENT_RETRIES and e.status_code in _TRANSIENT_STATUSES:
                time.sleep(2)
                continue
            logger.exception("flashcard agent HTTP error (status=%s)", e.status_code)
            raise
        except Exception:
            logger.exception("flashcard agent call failed")
            raise

    out: list[Card] = []
    for c in result.output.cards:
        front = (c.front or "").strip()
        back = (c.back or "").strip()
        if front and back:
            out.append({"front": front, "back": back})
    return out


def extract_cards_from_image(file_bytes: bytes, filename: str = "image.png") -> list[Card]:
    """OCR the image via the existing extraction pipeline, then ask the agent
    to split the markdown into Q/A pairs."""
    content_type = _guess_content_type(filename)
    markdown = extraction_service.extract_text_from_file(file_bytes, filename, content_type) or ""
    if not markdown.strip():
        return []
    prompt = _load_prompt("flashcard_ocr_split.txt").replace("{markdown}", markdown)
    return _run_flashcard_agent(prompt)


def gemini_generate_cards(source_text: str, count: int, difficulty: str) -> list[Card]:
    prompt = (
        _load_prompt("flashcard_generation.txt")
        .replace("{count}", str(count))
        .replace("{difficulty}", difficulty)
        .replace("{source}", source_text)
    )
    return _run_flashcard_agent(prompt)


def gemini_cleanup_cards(cards: list[Card]) -> list[Card]:
    prompt = _load_prompt("flashcard_cleanup.txt").replace(
        "{cards_json}", json.dumps(cards, ensure_ascii=False)
    )
    out = _run_flashcard_agent(prompt)
    return out if out else cards


def gemini_cloze(paragraph: str) -> list[Card]:
    prompt = _load_prompt("flashcard_cloze.txt").replace("{paragraph}", paragraph)
    return _run_flashcard_agent(prompt)


def generate_flashcards(
    topic: str,
    count: int = 5,
    context: str = "",
    documents: list[dict] | None = None,
    weak_concepts: list[str] | None = None,
) -> list[dict]:
    """Generate flashcards grounded in the student's course material, via the
    flashcard agent. Moved off gemini_service in #146.

    Args:
        topic:          The course or concept name.
        count:          Number of cards to generate.
        context:        Optional free-text context (e.g. session summary).
        documents:      Document dicts (file_name, category, summary, concept_notes).
        weak_concepts:  Concept names the student has low mastery on, weighted higher.
    """
    # ── Build document context block ──────────────────────────────────────────
    doc_blocks = []
    if documents:
        for doc in documents:
            parts = [f"[{doc.get('category', 'document').upper()}] {doc.get('file_name', '')}"]
            if doc.get("summary"):
                parts.append(f"Summary: {doc['summary']}")
            notes = doc.get("concept_notes")
            if notes and isinstance(notes, list):
                concept_lines = []
                for n in notes:
                    if not isinstance(n, dict):
                        continue
                    name = n.get("name")
                    desc = n.get("description")
                    if not name:
                        continue
                    concept_lines.append(f"- {name}: {desc}" if desc else f"- {name}")
                if concept_lines:
                    parts.append("Key concepts:\n" + "\n".join(concept_lines))
            doc_blocks.append("\n".join(parts))

    doc_context = ""
    if doc_blocks:
        doc_context = (
            "\n\nCOURSE MATERIAL (use this as the primary source for flashcard content):\n"
            + "\n\n---\n\n".join(doc_blocks)
        )

    # ── Weak concept focus block ──────────────────────────────────────────────
    weak_block = ""
    if weak_concepts:
        weak_block = (
            "\n\nThe student has LOW MASTERY on these concepts — prioritize them: "
            + ", ".join(weak_concepts)
        )

    # ── Free-text context (e.g. session summary) ──────────────────────────────
    extra_block = f"\n\nAdditional context:\n{context}" if context else ""

    prompt = f"""You are an expert tutor creating study flashcards for a student.

Course/Topic: "{topic}"{doc_context}{weak_block}{extra_block}

Generate exactly {count} flashcards.

Rules:
- Base card content on the course material provided above, not generic knowledge.
- Each card must have a clear FRONT (question or term) and a BACK (answer or definition).
- Vary difficulty: include recall, conceptual, and application questions.
- Prioritize concepts the student has low mastery on if listed above.
- Be specific — avoid vague or trivially obvious cards.
- Do NOT repeat questions already listed in the existing Q&A pairs above."""

    return _run_flashcard_agent(prompt)
